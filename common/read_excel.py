import openpyxl


def read_login_data():
    workbook = openpyxl.load_workbook("/Users/baiwenkai/PycharmProjects/pythonTestProject/mihuiAPI_Test_Pytest/"
                                      "mihuiapi_test_project/test_data/test_mihuiApi_excel.xlsx")
    sheet = workbook['login']

    allList = []
    for row in range(2, 7):
        tempList = []
        for column in range(1, 7):
            tempList.append(sheet.cell(row, column).value)
        allList.append(tempList)
    return allList


def read_register_data():
    workbook = openpyxl.load_workbook("/Users/baiwenkai/PycharmProjects/pythonTestProject/mihuiAPI_Test_Pytest/"
                                      "mihuiapi_test_project/test_data/test_mihuiApi_excel.xlsx")
    allList = []
    sheet = workbook['register']
    for row in range(3, sheet.max_row + 1):
        tempList = []
        for column in range(1, 9):
            tempList.append(sheet.cell(row, column).value)
        allList.append(tempList)
    return allList
print(read_register_data())